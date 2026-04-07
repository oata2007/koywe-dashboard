"""
data_loader.py
Carga y normaliza datos desde Google Sheets o archivos Excel para el Dashboard OTC de Koywe.
"""

from __future__ import annotations

import io
import json
import time
import unicodedata

import gspread
import pandas as pd
import streamlit as st
from datetime import datetime
from google.oauth2.service_account import Credentials

# ── IDs de Google Sheets por país ─────────────────────────────────────────────
SHEET_IDS = {
    "Chile":     "1UVWCq6sSXGVdc2wfS1gt257ny6nMancZLgBccdKcKbk",
    "Perú":      "12sh38s2GyUOVGrNYnr7eiYFENenobTzMdK0pobZg0CY",
    "México":    "18URXRrF31wMplQxQMKkGUyWS0QINTICJg6ZATOTFczA",
    "Brasil":    "1nBuw5Fk9cPPnqWQZRr6OKfSEgJFycOHO6jq035yYN58",
    "Argentina": "1v7Af32KPmKLUrXPQ24A_Ad8p-ZpdH69A5lPBEayLP0c",
    "USA":       "1fR8sUMFEPKjzBAQ4nAbKQKceV7iJEFtdCp-Am6KMRGs",
    "Colombia":  "115qtqs9JVB5m6skuiHTRXbBLooLeze4e9UDGY-MNEm4",
}

COUNTRY_CURRENCY = {
    "Chile": "CLP", "Perú": "PEN", "México": "MXN",
    "Brasil": "BRL", "Argentina": "ARS", "USA": "USD", "Colombia": "COP",
}

MONTHS_ES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]

QUARTERS = {
    "Q1 (Ene–Mar)": ["Enero", "Febrero", "Marzo"],
    "Q2 (Abr–Jun)": ["Abril", "Mayo", "Junio"],
    "Q3 (Jul–Sep)": ["Julio", "Agosto", "Septiembre"],
    "Q4 (Oct–Dic)": ["Octubre", "Noviembre", "Diciembre"],
}

# Tiers basados en volumen mensual USD (criterio OTC LatAm)
CLIENT_TIERS = [
    ("Enterprise", 5_000_000),   # > $5M
    ("Corporate",  1_000_000),   # $1M–$5M
    ("Business",     100_000),   # $100K–$1M
    ("Retail",             0),   # < $100K
]


# ── Auth ──────────────────────────────────────────────────────────────────────

def _get_gc():
    creds_dict = dict(st.secrets["gcp_service_account"])
    scopes = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    return gspread.authorize(creds)


def _drive_creds():
    """Retorna credenciales de Drive con token activo."""
    from google.auth.transport.requests import Request as GRequest
    creds_dict = dict(st.secrets["gcp_service_account"])
    creds = Credentials.from_service_account_info(
        creds_dict,
        scopes=["https://www.googleapis.com/auth/drive.readonly"],
    )
    creds.refresh(GRequest())
    return creds


def download_drive_file(file_id: str, dest_path: str) -> bool:
    """Descarga un archivo de Google Drive a dest_path. Retorna True si fue exitoso."""
    try:
        import requests as req
        creds = _drive_creds()
        url = f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media"
        r = req.get(url, headers={"Authorization": f"Bearer {creds.token}"},
                    stream=True, timeout=60)
        if r.status_code == 200:
            with open(dest_path, "wb") as f:
                for chunk in r.iter_content(chunk_size=32768):
                    f.write(chunk)
            return True
        return False
    except Exception:
        return False


def sync_drive_folder(folder_id: str,
                      dest_metrics: str = "/tmp/koywe_drive_metrics.xlsx",
                      dest_charts:  str = "/tmp/koywe_drive_charts.xlsx") -> dict:
    """
    Lista todos los .xlsx en la carpeta de Drive y los clasifica automáticamente:
      - Archivo con hoja 'Chart_1' → OTC Metrics
      - Archivo con hoja 'Chart_14' → OTC Charts
    Descarga los más recientes de cada tipo.
    Retorna {"metrics": bool, "charts": bool}.
    """
    import requests as req
    import openpyxl

    result = {"metrics": False, "charts": False}
    try:
        creds = _drive_creds()
        headers = {"Authorization": f"Bearer {creds.token}"}

        # Listar xlsx ordenados por fecha de modificación (más reciente primero)
        q = f"'{folder_id}' in parents and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' and trashed=false"
        list_url = "https://www.googleapis.com/drive/v3/files"
        resp = req.get(list_url, headers=headers,
                       params={"q": q, "orderBy": "modifiedTime desc",
                               "fields": "files(id,name,modifiedTime)", "pageSize": 20},
                       timeout=30)
        files = resp.json().get("files", [])

        for f in files:
            if result["metrics"] and result["charts"]:
                break  # Ya tenemos los dos, no seguir descargando

            # Descargar a un tmp temporal para inspeccionar
            _tmp = f"/tmp/_drive_inspect_{f['id']}.xlsx"
            ok = download_drive_file(f["id"], _tmp)
            if not ok:
                continue

            # Inspeccionar hojas y verificar que tengan datos reales
            try:
                wb = openpyxl.load_workbook(_tmp, read_only=True)
                sheets = wb.sheetnames

                def _has_data(ws) -> bool:
                    """True si la hoja tiene al menos una fila de datos (fila 3+)."""
                    for row in ws.iter_rows(min_row=3, max_row=5, values_only=True):
                        if any(v is not None for v in row):
                            return True
                    return False

                chart1_has_data  = "Chart_1"  in sheets and _has_data(wb["Chart_1"])
                chart14_has_data = "Chart_14" in sheets and _has_data(wb["Chart_14"])
                wb.close()
            except Exception:
                continue

            is_metrics = chart1_has_data  and not result["metrics"]
            is_charts  = chart14_has_data and not result["charts"]

            if is_metrics:
                import shutil
                shutil.copy(_tmp, dest_metrics)
                result["metrics"] = True

            if is_charts:
                import shutil
                shutil.copy(_tmp, dest_charts)
                result["charts"] = True

            try:
                import os as _os
                _os.remove(_tmp)
            except Exception:
                pass

    except Exception:
        pass

    return result


# ── Raw sheet loader ──────────────────────────────────────────────────────────

@st.cache_data(ttl=300, show_spinner=False)
def _load_raw(sheet_id: str, sheet_name: str) -> list[list]:
    """Lee valores crudos de una hoja. Cacheado 5 minutos. Reintenta ante 429."""
    for attempt in range(4):
        try:
            gc = _get_gc()
            ws = gc.open_by_key(sheet_id).worksheet(sheet_name)
            return ws.get_all_values()
        except gspread.exceptions.WorksheetNotFound:
            return []
        except gspread.exceptions.APIError as e:
            if "429" in str(e) and attempt < 3:
                time.sleep(2 ** attempt)   # 1s → 2s → 4s
                continue
            st.warning(f"No se pudo leer '{sheet_name}' ({sheet_id[:8]}…): {e}")
            return []
        except Exception as e:
            st.warning(f"No se pudo leer '{sheet_name}' ({sheet_id[:8]}…): {e}")
            return []
    return []


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_num(series: pd.Series) -> pd.Series:
    """
    Convierte texto numérico en formato US ($1,234.56 / 0.35%) a float.
    Trata '-' suelto como NaN.
    """
    cleaned = (
        series.astype(str)
              .str.strip()
              .str.replace(r"[$€\s]", "", regex=True)   # quita símbolo moneda
              .str.replace(",", "", regex=False)          # quita separador de miles
              .str.replace("%", "", regex=False)          # quita símbolo %
    )
    # '-' solo → NaN
    cleaned = cleaned.replace("-", float("nan"))
    return pd.to_numeric(cleaned, errors="coerce")


def _find_header_row(data: list[list]) -> int | None:
    """Encuentra la fila de headers (contiene 'Fecha' y 'Cliente' o 'Nombre del Cliente')."""
    for i, row in enumerate(data[:35]):
        has_fecha = "Fecha" in row
        has_cliente = "Cliente" in row or "Nombre del Cliente" in row
        if has_fecha and has_cliente:
            return i
    return None


def _dedup_headers(headers: list[str]) -> list[str]:
    """Deduplica nombres de columna vacíos o repetidos."""
    seen: dict[str, int] = {}
    result = []
    for h in headers:
        h = h.strip() or "_col"
        count = seen.get(h, 0)
        result.append(h if count == 0 else f"{h}_{count}")
        seen[h] = count + 1
    return result


# Palabras clave para excluir registros internos (case-insensitive)
EXCLUDED_KEYWORDS = ["off colombia", "traspaso mbi", "cobertura rampa", "webon", "web on"]

# ── Directorio de clientes (email → nombre) ───────────────────────────────────
_CLIENT_NAMES_PATH    = "data/client_names.xlsx"       # Base Tomas OTC
_CLIENT_EXTRA_PATH    = "data/clientes_extra.xlsx"     # clientes_sin_nombre completado

def _clean_email(val) -> str | None:
    if pd.isna(val):
        return None
    s = str(val).strip().lower()
    return s if "@" in s and "." in s else None

def _clean_name(val) -> str | None:
    if pd.isna(val):
        return None
    s = str(val).strip()
    skip = {"clientes", "cliente", "nombre", "nan", "", "email", "mail"}
    return s if s.lower() not in skip else None

@st.cache_data(show_spinner=False)
def load_client_names() -> dict[str, str]:
    """
    Combina el mapeo email→nombre de dos fuentes:
    1. data/client_names.xlsx  (Base Tomas OTC — hojas CHILE/PERU/MEXICO/RECUENTO)
    2. data/clientes_extra.xlsx (clientes_sin_nombre completado — hoja 'Sin Nombre')
    La segunda fuente sobreescribe a la primera si hay conflicto.
    """
    mapping: dict[str, str] = {}

    # ── Fuente 1: Base Tomas OTC ──────────────────────────────────────────────
    try:
        xl1 = pd.ExcelFile(_CLIENT_NAMES_PATH)
        for sheet, hdr in [("CHILE", 1), ("PERU", 1), ("MEXICO", 0),
                            ("ARGENTINA", 1), ("RECUENTO", 0)]:
            if sheet not in xl1.sheet_names:
                continue
            try:
                df = pd.read_excel(_CLIENT_NAMES_PATH, sheet_name=sheet, header=hdr)
                # Tomar TODAS las columnas de nombre y email (puede haber múltiples)
                name_cols  = [c for c in df.columns if "CLIENTE" in str(c).upper()
                              or "NOMBRE" in str(c).upper()]
                email_cols = [c for c in df.columns if "MAIL" in str(c).upper()
                              or "EMAIL" in str(c).upper()]
                for nc in name_cols:
                    for ec in email_cols:
                        for _, row in df.iterrows():
                            n, e = _clean_name(row[nc]), _clean_email(row[ec])
                            if n and e and e not in mapping:
                                mapping[e] = n
            except Exception:
                continue
    except Exception:
        pass

    # ── Fuente 2: clientes_extra (sobreescribe) ───────────────────────────────
    try:
        xl2    = pd.ExcelFile(_CLIENT_EXTRA_PATH)
        sheet2 = "Sin Nombre" if "Sin Nombre" in xl2.sheet_names else xl2.sheet_names[0]
        df2    = pd.read_excel(_CLIENT_EXTRA_PATH, sheet_name=sheet2, header=0)
        email_col = next(
            (c for c in df2.columns if "EMAIL" in str(c).upper()
             or "MAIL" in str(c).upper()), None
        )
        name_col = next(
            (c for c in df2.columns if "NOMBRE" in str(c).upper()
             or "NOMBRE" in str(c).upper()), None
        )
        if email_col and name_col:
            for _, row in df2.iterrows():
                n, e = _clean_name(row[name_col]), _clean_email(row[email_col])
                if n and e:
                    mapping[e] = n   # sobreescribe fuente 1
    except Exception:
        pass

    return mapping


# ── Normalize ─────────────────────────────────────────────────────────────────

def _normalize(df: pd.DataFrame, country: str, sheet_name: str) -> pd.DataFrame:
    df = df.copy()
    df.columns = [c.strip() for c in df.columns]
    # Normalizar nombre de columna cliente (Perú/México usan "Nombre del Cliente")
    if "Nombre del Cliente" in df.columns and "Cliente" not in df.columns:
        df = df.rename(columns={"Nombre del Cliente": "Cliente"})

    # Fecha — formato M/D/YYYY (ej. 3/2/2026)
    df["Fecha"] = pd.to_datetime(df["Fecha"], dayfirst=False, errors="coerce")
    df = df[df["Fecha"].notna()].copy()

    # Excluir registros internos (case-insensitive) — revisa todas las columnas de texto
    str_cols = [c for c in df.columns if df[c].dtype == object]
    combined = df[str_cols].fillna("").astype(str).apply(
        lambda row: " ".join(row.values).lower(), axis=1
    )
    mask = combined.apply(lambda v: any(kw in v for kw in EXCLUDED_KEYWORDS))
    df = df[~mask].copy()
    if df.empty:
        return df

    # Columnas numéricas estándar
    for col in ["TC Cliente", "Monto USDT", "Monto USD", "Spread", "% Revenue", "Gross Revenue"]:
        if col in df.columns:
            df[col] = _parse_num(df[col])

    # Columna de moneda local (Monto CLP, Monto MXN, etc.)
    local_code = COUNTRY_CURRENCY.get(country, "")
    local_col  = f"Monto {local_code}"
    if local_col in df.columns:
        df[local_col] = _parse_num(df[local_col])
        df["Monto Local"] = df[local_col]
    else:
        # Detección automática si el nombre varía
        for c in df.columns:
            if c.startswith("Monto ") and c not in ("Monto USDT", "Monto USD"):
                df[c] = _parse_num(df[c])
                df["Monto Local"] = df[c]
                break

    df["Moneda"] = local_code
    df["País"]   = country
    df["Hoja"]   = sheet_name  # "Enero 2026", etc.

    # Periodo helpers
    df["Año"]    = df["Fecha"].dt.year
    df["Mes_n"]  = df["Fecha"].dt.month
    df["Semana"] = df["Fecha"].dt.to_period("W").astype(str)

    return df


# ── Mapeo de códigos de país (Excel unificado) ────────────────────────────────

COUNTRY_MAP: dict[str, str] = {
    "CHL": "Chile",
    "PER": "Perú",
    "MEX": "México",
    "BRA": "Brasil",
    "ARG": "Argentina",
    "USA": "USA",
    "COL": "Colombia",
}


def _parse_channel(val) -> str:
    """Convierte la columna metadata al canal legible."""
    if pd.isna(val):
        return "Manual"
    s = str(val).strip()
    try:
        data = json.loads(s)
        s = data.get("oldMetadata", s)
    except Exception:
        pass
    sl = s.lower()
    if "web-otc" in sl:
        return "Web OTC"
    if "k3" in sl:
        return "K3"
    return "Manual"


# ── Country detection from filename ───────────────────────────────────────────

def _strip_accents(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )


# Alias para detectar país desde el nombre del archivo
_COUNTRY_ALIASES: dict[str, str] = {
    "chile": "Chile",
    "peru": "Perú", "perú": "Perú",
    "mexico": "México", "méxico": "México",
    "brasil": "Brasil", "brazil": "Brasil",
    "argentina": "Argentina",
    "usa": "USA", "estados unidos": "USA", "eeuu": "USA",
    "colombia": "Colombia",
}


def detect_country_from_filename(filename: str) -> str | None:
    """Infiere el país a partir del nombre del archivo (sin extensión, sin tildes)."""
    stem = filename.rsplit(".", 1)[0].strip().lower()
    stem_clean = _strip_accents(stem)
    return (
        _COUNTRY_ALIASES.get(stem)
        or _COUNTRY_ALIASES.get(stem_clean)
        or next(
            (v for k, v in _COUNTRY_ALIASES.items() if k in stem_clean),
            None,
        )
    )


# ── Excel loader ───────────────────────────────────────────────────────────────

def _load_raw_excel(file_bytes: bytes, sheet_name: str) -> list[list]:
    """Lee valores crudos de una hoja de un xlsx en memoria."""
    try:
        df = pd.read_excel(
            io.BytesIO(file_bytes),
            sheet_name=sheet_name,
            header=None,
            dtype=str,
        )
        df = df.fillna("")
        return df.values.tolist()
    except Exception:
        return []


def load_month_excel(
    country: str, month: str, year: int, file_bytes: bytes
) -> pd.DataFrame:
    """Carga un mes de un país desde un archivo Excel en memoria."""
    sheet_name = f"{month} {year}"
    raw = _load_raw_excel(file_bytes, sheet_name)
    if not raw:
        return pd.DataFrame()
    hi = _find_header_row(raw)
    if hi is None:
        return pd.DataFrame()
    headers = _dedup_headers(raw[hi])
    rows = raw[hi + 1:]
    df = pd.DataFrame(rows, columns=headers)
    df = df[df["Fecha"].str.strip() != ""].copy()
    return _normalize(df, country, sheet_name)


def load_all_excel(
    months: tuple, year: int, uploaded: dict[str, bytes]
) -> pd.DataFrame:
    """
    Carga todos los países desde archivos Excel subidos.
    `uploaded` es {country: file_bytes}.
    Sin cache: los bytes ya están en memoria.
    """
    dfs = []
    for country, file_bytes in uploaded.items():
        for month in months:
            df = load_month_excel(country, month, year, file_bytes)
            if not df.empty:
                dfs.append(df)
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()


# ── Loader archivo Excel unificado (Chart_14) ─────────────────────────────────

def load_from_single_excel(
    file_bytes: bytes, months_tuple: tuple, year: int
) -> pd.DataFrame:
    """
    Carga datos del archivo Excel unificado (hoja Chart_14).

    Fórmulas OTC:
      FX ajustado  = exchangeRate / 0.9995
      Monto USDT   = monto / FX_ajustado
      Gross Revenue = (monto / hedge.unitPrice) - Monto USDT

    Columnas de entrada:
      A creación | B Usuario | D exchangeRate | E monto
      K countryCode | L metadata | M hedge.unitPrice
    """
    try:
        df = pd.read_excel(
            io.BytesIO(file_bytes),
            sheet_name="Chart_14",
            header=1,           # encabezado en fila 2 de Excel
        )
    except Exception as e:
        st.warning(f"Error al leer Chart_14: {e}")
        return pd.DataFrame()

    df = df.rename(columns={
        "creación":        "Fecha",
        "Usuario":         "Cliente",
        "exchangeRate":    "_fx",
        "monto":           "_monto",
        "countryCode":     "_country",
        "metadata":        "_meta",
        "hedge.unitPrice": "_hedge",
    })

    missing = {"Fecha", "Cliente", "_fx", "_monto", "_country", "_hedge"} - set(df.columns)
    if missing:
        st.warning(f"Faltan columnas en Chart_14: {missing}")
        return pd.DataFrame()

    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df = df[df["Fecha"].notna()].copy()

    # Filtrar por año y meses solicitados
    month_nums = tuple(
        MONTHS_ES.index(m) + 1 for m in months_tuple if m in MONTHS_ES
    )
    df = df[df["Fecha"].dt.year == year]
    df = df[df["Fecha"].dt.month.isin(month_nums)]
    if df.empty:
        return pd.DataFrame()

    # Métricas OTC
    adj_fx            = df["_fx"] / 0.9995
    df["Monto USDT"]  = df["_monto"] / adj_fx
    df["Monto USD"]   = df["Monto USDT"]
    df["Gross Revenue"] = (df["_monto"] / df["_hedge"]) - df["Monto USDT"]

    # Mapeos
    df["País"]  = df["_country"].map(COUNTRY_MAP).fillna(df["_country"])
    df["Canal"] = df["_meta"].apply(_parse_channel)

    # Solo incluir canales web-otc y k3 (filtro previo a mapeo)
    df = df[df["_meta"].astype(str).str.lower().str.contains("web-otc|koywe-k3", na=False)].copy()
    if df.empty:
        return pd.DataFrame()

    # Exclusiones internas
    excl = df["Cliente"].astype(str).str.lower().apply(
        lambda v: any(kw in v for kw in EXCLUDED_KEYWORDS)
    )
    df = df[~excl].copy()

    # Columnas auxiliares (compatibilidad con in_period y client_analysis)
    df["Hoja"]       = df["Fecha"].dt.month.apply(lambda m: f"{MONTHS_ES[m - 1]} {year}")
    df["Año"]        = year
    df["Mes_n"]      = df["Fecha"].dt.month
    df["Semana"]     = df["Fecha"].dt.to_period("W").astype(str)
    df["Moneda"]     = df["_country"].map({
        "CHL": "CLP", "PER": "PEN", "MEX": "MXN",
        "BRA": "BRL", "ARG": "ARS", "USA": "USD", "COL": "COP",
    }).fillna("")
    df["Monto Local"] = df["_monto"]

    keep = [
        "Fecha", "Cliente", "País", "Monto USD", "Monto USDT", "Gross Revenue",
        "Canal", "Hoja", "Año", "Mes_n", "Semana", "Moneda", "Monto Local",
    ]
    return df[[c for c in keep if c in df.columns]].reset_index(drop=True)


# ── Public API (Google Sheets) ─────────────────────────────────────────────────

def load_month(country: str, month: str, year: int) -> pd.DataFrame:
    """Carga un mes de un país. Retorna DataFrame normalizado o vacío."""
    sheet_id   = SHEET_IDS.get(country)
    if not sheet_id:
        return pd.DataFrame()
    sheet_name = f"{month} {year}"
    raw = _load_raw(sheet_id, sheet_name)
    if not raw:
        return pd.DataFrame()

    hi = _find_header_row(raw)
    if hi is None:
        return pd.DataFrame()

    headers = _dedup_headers(raw[hi])
    rows    = raw[hi + 1:]
    df      = pd.DataFrame(rows, columns=headers)

    # Descartar filas sin fecha
    df = df[df["Fecha"].str.strip() != ""].copy()
    df = df[df["Fecha"].notna()].copy()

    return _normalize(df, country, sheet_name)


@st.cache_data(ttl=300, show_spinner=False)
def load_all(months: tuple[str, ...], year: int) -> pd.DataFrame:
    """
    Carga todos los países para los meses indicados.
    `months` es tuple (hashable) para que funcione el cache.
    """
    dfs = []
    for country in SHEET_IDS:
        for month in months:
            df = load_month(country, month, year)
            if not df.empty:
                dfs.append(df)
            time.sleep(0.4)   # evitar quota 429 en primera carga
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()


# ── Period helpers ────────────────────────────────────────────────────────────

def months_for_period(view: str, selected: str, year: int) -> tuple[list[str], list[str]]:
    """
    Retorna (meses_actuales, meses_anteriores) según la vista elegida.
    Los meses son solo el nombre: "Enero", "Febrero", etc.
    """
    today = datetime.now()
    available = MONTHS_ES[:today.month] if year == today.year else MONTHS_ES

    if view == "Mensual":
        idx     = MONTHS_ES.index(selected)
        current = [MONTHS_ES[idx]]
        prev    = [MONTHS_ES[idx - 1]] if idx > 0 else []

    elif view == "Trimestral":
        q_months = QUARTERS[selected]
        current  = [m for m in q_months if m in available]
        # Trimestre anterior
        q_keys = list(QUARTERS.keys())
        q_idx  = q_keys.index(selected)
        prev   = list(QUARTERS[q_keys[q_idx - 1]]) if q_idx > 0 else []

    else:  # YTD
        current = list(available)
        prev    = []

    return current, prev


def available_months(year: int) -> list[str]:
    today = datetime.now()
    return MONTHS_ES[:today.month] if year == today.year else MONTHS_ES


def available_quarters(year: int) -> list[str]:
    avail = available_months(year)
    return [q for q, ms in QUARTERS.items() if any(m in avail for m in ms)]


# ── Analytics ─────────────────────────────────────────────────────────────────

def assign_tier(vol_usd: float) -> str:
    for name, threshold in CLIENT_TIERS:
        if vol_usd >= threshold:
            return name
    return "Retail"


def client_analysis(df_curr: pd.DataFrame, df_prev: pd.DataFrame) -> dict:
    """
    Calcula clientes nuevos, recurrentes y perdidos.
    También genera tabla de clientes con tier, volumen y revenue.
    """
    def client_set(df):
        if df.empty or "Cliente" not in df.columns:
            return set()
        return set(df["Cliente"].dropna().str.strip().unique())

    curr_clients = client_set(df_curr)
    prev_clients = client_set(df_prev)

    new       = curr_clients - prev_clients
    recurring = curr_clients & prev_clients
    lost      = prev_clients - curr_clients

    # Tabla de clientes (período actual)
    client_df = pd.DataFrame()
    if not df_curr.empty and "Cliente" in df_curr.columns:
        agg = df_curr.groupby("Cliente").agg(
            Volumen_USD   = ("Monto USD",      "sum"),
            Revenue_USD   = ("Gross Revenue",  "sum"),
            USDT_USD      = ("Monto USDT",     "sum"),
            Operaciones   = ("Fecha",          "count"),
            Países        = ("País",           lambda x: ", ".join(sorted(x.unique()))),
        ).reset_index()

        agg["Takerate_%"] = (agg["Revenue_USD"] / agg["USDT_USD"] * 100).round(4)
        agg["Tier"]       = agg["Volumen_USD"].apply(assign_tier)
        agg["Estado"]     = agg["Cliente"].apply(
            lambda c: "Nuevo" if c in new else ("Recurrente" if c in recurring else "Activo")
        )
        client_df = agg.sort_values("Volumen_USD", ascending=False)

    return {
        "new": sorted(new),
        "recurring": sorted(recurring),
        "lost": sorted(lost),
        "n_new": len(new),
        "n_recurring": len(recurring),
        "n_lost": len(lost),
        "client_df": client_df,
    }


# ── Excel dashboard loader (Mongo export: Chart_1/2/4/11) ─────────────────────

COUNTRY_MAP_CODE = {
    "CHL": "Chile", "PER": "Perú", "MEX": "México",
    "BRA": "Brasil", "ARG": "Argentina", "USA": "USA", "COL": "Colombia",
}

def load_excel_dashboard(file_bytes: bytes) -> dict:
    """Load aggregated dashboard data from Mongo Excel export (Chart_1/2/4/11)."""
    def _read(buf, sheet, cols):
        buf.seek(0)
        try:
            df = pd.read_excel(buf, sheet_name=sheet, header=1)
        except Exception:
            return pd.DataFrame(columns=cols)
        if df.empty or len(df.columns) < len(cols):
            return pd.DataFrame(columns=cols)
        # Rename columns by position
        rename = {old: new for old, new in zip(df.columns, cols)}
        df = df.rename(columns=rename)
        # Verificar que las columnas esperadas existen
        for c in cols:
            if c not in df.columns:
                return pd.DataFrame(columns=cols)
        return df

    buf = io.BytesIO(file_bytes)

    def _parse_periodo(series: pd.Series) -> pd.Series:
        """Parsea fechas en múltiples formatos: 'Feb-2026', 'Feb 2026', '2026-02', datetime, etc."""
        # Si ya son datetime (openpyxl las parsea automáticamente en Python 3.14)
        if pd.api.types.is_datetime64_any_dtype(series):
            return series.dt.to_period("M").dt.to_timestamp()
        # Si los valores son objetos datetime de Python
        sample = series.dropna()
        if not sample.empty and hasattr(sample.iloc[0], "year"):
            return pd.to_datetime(series, errors="coerce").dt.to_period("M").dt.to_timestamp()
        # Intentar múltiples formatos de texto
        for fmt in ["%b-%Y", "%b %Y", "%Y-%m", "%m-%Y", "%B-%Y", "%B %Y"]:
            parsed = pd.to_datetime(series, format=fmt, errors="coerce")
            if parsed.notna().sum() > 0:
                return parsed
        # Último recurso: inferir formato automáticamente
        return pd.to_datetime(series, errors="coerce", dayfirst=False)

    # Chart_1: Volume USDT by country/month
    df1 = _read(buf, "Chart_1", ["Volumen_USDT", "Periodo", "countryCode"])
    if not df1.empty:
        df1["Fecha"] = _parse_periodo(df1["Periodo"])
        df1["País"] = df1["countryCode"].map(COUNTRY_MAP_CODE).fillna(df1["countryCode"])
        df1["Mes"] = df1["Fecha"].dt.month
        df1["Año"] = df1["Fecha"].dt.year
        df1 = df1.dropna(subset=["Fecha", "Volumen_USDT"]).copy()
        df1["Volumen_USDT"] = pd.to_numeric(df1["Volumen_USDT"], errors="coerce").fillna(0)

    # Chart_2: Volume by client/month
    df2 = _read(buf, "Chart_2", ["Volumen_USD", "Periodo", "Cliente"])
    if not df2.empty:
        df2["Fecha"] = _parse_periodo(df2["Periodo"])
        df2["Mes"] = df2["Fecha"].dt.month
        df2["Año"] = df2["Fecha"].dt.year
        df2 = df2.dropna(subset=["Fecha", "Volumen_USD"]).copy()
        df2["Volumen_USD"] = pd.to_numeric(df2["Volumen_USD"], errors="coerce").fillna(0)

    # Chart_4: Spread by client/month
    df4 = _read(buf, "Chart_4", ["Spread", "Cliente", "Periodo"])
    df4["Spread"] = pd.to_numeric(df4["Spread"], errors="coerce")
    df4 = df4.dropna(subset=["Spread"]).copy()

    # Join clients + spread → Revenue
    df_clients = df2.merge(
        df4[["Cliente", "Periodo", "Spread"]],
        on=["Cliente", "Periodo"],
        how="left",
    )
    df_clients["Spread"]      = pd.to_numeric(df_clients["Spread"],      errors="coerce").fillna(0)
    df_clients["Volumen_USD"] = pd.to_numeric(df_clients["Volumen_USD"], errors="coerce").fillna(0)
    df_clients["Revenue"]     = (df_clients["Volumen_USD"] * df_clients["Spread"]).round(2)
    df_clients["Takerate_pct"] = (df_clients["Spread"] * 100).round(4)

    # Chart_11: Daily volume by country
    df11 = _read(buf, "Chart_11", ["Fecha_str", "Volumen_local", "countryCode"])
    if not df11.empty:
        df11["Fecha"] = pd.to_datetime(df11["Fecha_str"], format="%d-%b-%Y", errors="coerce")
        df11["País"] = df11["countryCode"].map(COUNTRY_MAP_CODE).fillna(df11["countryCode"])
        df11["Volumen_local"] = pd.to_numeric(df11["Volumen_local"], errors="coerce").fillna(0)
        df11 = df11.dropna(subset=["Fecha"]).copy()

    # Asegurar que todos los DataFrames tienen las columnas esperadas aunque estén vacíos
    def _safe_select(df, cols):
        for c in cols:
            if c not in df.columns:
                df[c] = pd.NA
        return df[cols].reset_index(drop=True)

    return {
        "vol_country": _safe_select(df1, ["País", "Periodo", "Fecha", "Mes", "Año", "Volumen_USDT"]),
        "clients":     _safe_select(df_clients, ["Cliente", "Periodo", "Fecha", "Mes", "Año", "Volumen_USD", "Spread", "Revenue", "Takerate_pct"]),
        "daily":       _safe_select(df11, ["Fecha", "País", "Volumen_local"]),
    }


def client_analysis_excel(df_clients: pd.DataFrame, curr_months: list, curr_year: int,
                           prev_months: list | None = None, prev_year: int | None = None) -> dict:
    """Compute new/recurring/lost/reactivated clients from aggregated Excel client data."""
    curr_mask = (df_clients["Año"] == curr_year) & (df_clients["Mes"].isin(curr_months))
    curr_set = set(df_clients.loc[curr_mask, "Cliente"].unique())

    if prev_months and prev_year is not None:
        prev_mask = (df_clients["Año"] == prev_year) & (df_clients["Mes"].isin(prev_months))
    elif prev_months:
        prev_mask = (df_clients["Año"] == curr_year) & (df_clients["Mes"].isin(prev_months))
    else:
        prev_mask = pd.Series(False, index=df_clients.index)
    prev_set = set(df_clients.loc[prev_mask, "Cliente"].unique())

    # All clients before current period start
    if not df_clients.loc[curr_mask].empty:
        cutoff = df_clients.loc[curr_mask, "Fecha"].min()
        hist_set = set(df_clients.loc[df_clients["Fecha"] < cutoff, "Cliente"].unique())
    else:
        hist_set = set()

    new_set        = curr_set - hist_set
    recurring_set  = curr_set & prev_set
    lost_set       = prev_set - curr_set
    reactivated_set = (curr_set & (hist_set - prev_set)) - new_set

    # Client summary table
    curr_data = df_clients.loc[curr_mask].groupby("Cliente").agg(
        Volumen_USD=("Volumen_USD", "sum"),
        Revenue=("Revenue", "sum"),
        Meses=("Periodo", "nunique"),
    ).reset_index()
    curr_data["Takerate_%"] = (curr_data["Revenue"] / curr_data["Volumen_USD"].replace(0, float("nan")) * 100).round(4)

    def _tier(v):
        if v >= 5_000_000: return "Enterprise"
        if v >= 1_000_000: return "Corporate"
        if v >= 100_000:   return "Business"
        return "Retail"
    curr_data["Tier"] = curr_data["Volumen_USD"].apply(_tier)

    def _estado(c):
        if c in new_set:        return "Nuevo"
        if c in reactivated_set: return "Reactivado"
        if c in recurring_set:  return "Recurrente"
        return "Activo"
    curr_data["Estado"] = curr_data["Cliente"].apply(_estado)

    return {
        "new": new_set,
        "recurring": recurring_set,
        "lost": lost_set,
        "reactivated": reactivated_set,
        "n_new": len(new_set),
        "n_recurring": len(recurring_set),
        "n_lost": len(lost_set),
        "n_reactivated": len(reactivated_set),
        "client_df": curr_data,
    }


# ── Weekly summary (Chart_14 transaction data) ────────────────────────────────
@st.cache_data(show_spinner=False)
def load_weekly_summary(file_path: str, year: int) -> pd.DataFrame:
    """Carga Chart_14 y agrega por semana ISO para el año dado."""
    try:
        df = pd.read_excel(file_path, sheet_name="Chart_14", header=1)
    except Exception as e:
        return pd.DataFrame()

    df["creación"] = pd.to_datetime(df["creación"], errors="coerce")
    df = df[df["creación"].dt.year == year].copy()
    if df.empty:
        return pd.DataFrame()

    df["adj_fx"]     = df["exchangeRate"] / 0.9995
    df["Vol_USDT"]   = df["monto"] / df["adj_fx"]
    df["Revenue"]    = (df["monto"] / df["hedge.unitPrice"]) - df["Vol_USDT"]
    df["País"]       = df["countryCode"].map(COUNTRY_MAP_CODE).fillna(df["countryCode"])

    def _canal(v):
        s = str(v).lower()
        try:
            s = json.loads(v).get("oldMetadata", s).lower()
        except Exception:
            pass
        if "web-otc" in s:  return "Web OTC"
        if "k3" in s:       return "K3"
        return "Manual"

    df["Canal"] = df["metadata"].apply(_canal)

    iso          = df["creación"].dt.isocalendar()
    df["W_num"]  = iso["week"].astype(int)
    df["W_start"] = df["creación"].dt.to_period("W").apply(lambda p: p.start_time)

    grp = df.groupby(["W_num", "W_start"])
    weekly = grp.agg(
        Volumen_USDT=("Vol_USDT", "sum"),
        Operaciones =("Vol_USDT", "count"),
        Revenue_USD =("Revenue",  "sum"),
        Clientes    =("Usuario",  "nunique"),
    ).reset_index().sort_values("W_start")

    # Top canal, país y etiqueta con fechas reales de transacciones
    rows = []
    seq  = 1  # número secuencial de semana dentro del año
    for _, row in weekly.iterrows():
        mask = (df["W_num"] == row["W_num"]) & (df["W_start"] == row["W_start"])
        wdf  = df[mask]
        top_canal  = wdf.groupby("Canal")["Vol_USDT"].sum().idxmax() if len(wdf) else "—"
        top_pais   = wdf.groupby("País")["Vol_USDT"].sum().idxmax()  if len(wdf) else "—"
        # Usar rango real de fechas dentro de esa semana
        real_start = wdf["creación"].min().normalize()
        real_end   = wdf["creación"].max().normalize()
        if real_start == real_end:
            label = f"W{seq}  {real_start.strftime('%d %b')}"
        else:
            label = f"W{seq}  {real_start.strftime('%d %b')} – {real_end.strftime('%d %b')}"
        rows.append({**row.to_dict(), "Semana": label,
                     "Top Canal": top_canal, "Top País": top_pais})
        seq += 1

    return pd.DataFrame(rows)


@st.cache_data(show_spinner=False)
def load_canal_monthly(file_path: str) -> pd.DataFrame:
    """Carga Chart_14 y agrega volumen USDT por canal y mes (todos los años)."""
    try:
        df = pd.read_excel(file_path, sheet_name="Chart_14", header=1)
    except Exception:
        return pd.DataFrame()

    df["creación"] = pd.to_datetime(df["creación"], errors="coerce")
    df = df[df["creación"].notna()].copy()
    if df.empty:
        return pd.DataFrame()

    df["adj_fx"]   = df["exchangeRate"] / 0.9995
    df["Vol_USDT"] = df["monto"] / df["adj_fx"]

    def _canal(v):
        s = str(v).lower()
        try:
            s = json.loads(v).get("oldMetadata", s).lower()
        except Exception:
            pass
        if "web-otc" in s: return "Web OTC"
        if "k3" in s:      return "K3"
        return "Manual"

    df["Canal"] = df["metadata"].apply(_canal)
    df["Fecha"] = df["creación"].dt.to_period("M").dt.to_timestamp()
    df["Mes"]   = df["creación"].dt.month
    df["Año"]   = df["creación"].dt.year

    return (
        df.groupby(["Canal", "Fecha", "Mes", "Año"])["Vol_USDT"]
        .sum()
        .reset_index()
        .rename(columns={"Vol_USDT": "Volumen_USDT"})
        .sort_values("Fecha")
    )
